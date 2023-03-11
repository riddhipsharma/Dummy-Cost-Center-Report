from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
# 1 -> lastmonth sheet
# 2 -> currentmonth sheet
def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        # print ("Hello World") 
        excel_file1 = request.FILES["excel_file1"]  
        excel_file2 = request.FILES["excel_file2"]
        wb1 = load_workbook(excel_file1)
        wb2 = load_workbook(excel_file2)
        # wb = Workbook()
        # ws = wb.create_sheet("Sheet1",0)
        ws1 = wb1['Sheet1']
        ws2 = wb2['Sheet1']

        ws2.insert_cols(idx=3)
        char = get_column_letter(3)
        ws2[char + str(1)].value = "ACTUAL_COST_CENTER"
        i=1
        # col is A B C D ...
        for row1 in range(2,201):
            for col1 in range(4,5):
                # print(i)
                char1 = get_column_letter(col1)
                # print(ws1[char1 + str(row1)])
                s1=ws1[char1 + str(row1)].value
                for row4 in range(2,201):
                    for col4 in range(4,5):
                        char4 = get_column_letter(col4)
                        # print(ws2[char4 + str(row4)])
                        s2=ws2[char4 + str(row4)].value
                        if(s1==s2):
                            # print("same sample number")

                            for col2 in range(8,9):
                                char2 = get_column_letter(col2)
                                # print(ws1[char2 + str(row1)])
                                # print(ws2[char2 + str(row4)])
                                t1=ws1[char2 + str(row1)].value
                                t2=ws2[char2 + str(row4)].value
                                if t1==t2:
                                    # print("same test code")
                                    for col3 in range(3,4):
                                        char3 = get_column_letter(col3)
                                        # print(ws2[char3 + str(row1)])
                                        # print("charge code updated")
                                        # print()
                                        # ws[char3 + str(row4)].value=ws1[char3 + str(row1)].value
                                        ws2[char3 + str(row4)].value=ws1[char3 + str(row1)].value
                                        # wb2.save("excel_file2")
                                
                        else:
                            # print("Different sample number")
                            continue             
                i=i+1
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=DummyCostCenter.xlsx'
        wb2.save(response)
        return response
